"""
Finance chart generator — 월별 수입/지출 추이 그래프 및 Excel 파일 생성.
"""
import io
import json
import logging
from collections import defaultdict
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

_FINANCE_DIR = Path(__file__).resolve().parents[2] / "data" / "finance"
_SNAPSHOT_PATH = _FINANCE_DIR / "monthly_snapshot.json"
_TRANSACTIONS_PATH = _FINANCE_DIR / "transactions.json"
_FINANCE_ASSETS_PATH = _FINANCE_DIR / "finance_assets.json"


# ── 한글 폰트 설정 ────────────────────────────────────────────────────────────

def _setup_korean_font():
    candidates = ["AppleGothic", "Apple SD Gothic Neo", "NanumGothic", "Malgun Gothic", "NanumBarunGothic"]
    available = {f.name for f in fm.fontManager.ttflist}
    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            return
    plt.rcParams["axes.unicode_minus"] = False


# ── 데이터 로드 헬퍼 ──────────────────────────────────────────────────────────

def _load_snapshots() -> list[dict]:
    if not _SNAPSHOT_PATH.exists():
        return []
    try:
        return json.loads(_SNAPSHOT_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def _load_transactions() -> list[dict]:
    if not _TRANSACTIONS_PATH.exists():
        return []
    try:
        return json.loads(_TRANSACTIONS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def _load_finance_assets() -> dict:
    if not _FINANCE_ASSETS_PATH.exists():
        return {"fixed": {"salary": 0, "savings": 0, "fixed_expenses": []}}
    try:
        return json.loads(_FINANCE_ASSETS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"fixed": {"salary": 0, "savings": 0, "fixed_expenses": []}}


def _monthly_income(month: str, txs: list[dict]) -> int:
    return sum(t["amount"] for t in txs if t.get("date", "")[:7] == month and t.get("source") == "income")


def _monthly_card_breakdown(month: str, txs: list[dict]) -> dict[str, int]:
    bd: dict[str, int] = defaultdict(int)
    for t in txs:
        if t.get("date", "")[:7] == month and t.get("source") not in ("monthly_total", "unknown", "income"):
            bd[t.get("card") or "기타"] += t.get("amount", 0)
    return dict(bd)


def _build_monthly_data(months: int = 6) -> list[dict]:
    """최근 N개월 수입/지출 데이터 구성."""
    snaps = sorted(_load_snapshots(), key=lambda s: s["month"])[-months:]
    txs = _load_transactions()
    fa = _load_finance_assets()
    fixed = fa.get("fixed", {})
    salary = fixed.get("salary", 0)
    savings_monthly = fixed.get("savings", 0)
    fixed_exp_total = sum(e.get("amount", 0) for e in fixed.get("fixed_expenses", [])) + savings_monthly

    rows = []
    for snap in snaps:
        m = snap["month"]
        extra = _monthly_income(m, txs)
        card_spend = snap.get("card_spend", 0)
        total_income = salary + extra
        total_expense = fixed_exp_total + card_spend
        remainder = total_income - total_expense
        rows.append({
            "month": m,
            "salary": salary,
            "extra_income": extra,
            "total_income": total_income,
            "fixed_expense": fixed_exp_total,
            "card_spend": card_spend,
            "total_expense": total_expense,
            "remainder": remainder,
            "card_breakdown": _monthly_card_breakdown(m, txs),
        })
    return rows


# ── 차트 생성 ─────────────────────────────────────────────────────────────────

def generate_chart(months: int = 4) -> bytes:
    """최근 N개월 수입/지출 바 차트 PNG bytes 반환."""
    _setup_korean_font()
    data = _build_monthly_data(months)
    if not data:
        raise ValueError("스냅샷 데이터 없음")

    labels = [d["month"][5:] + "월" for d in data]  # "2026-04" → "04월"
    incomes = [d["total_income"] / 10000 for d in data]
    expenses = [d["total_expense"] / 10000 for d in data]
    remainders = [d["remainder"] / 10000 for d in data]

    x = range(len(labels))
    width = 0.35

    fig, ax1 = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor("#1e1e2e")
    ax1.set_facecolor("#1e1e2e")

    bars_income = ax1.bar([i - width / 2 for i in x], incomes, width, label="수입", color="#4ade80", alpha=0.85)
    bars_expense = ax1.bar([i + width / 2 for i in x], expenses, width, label="지출", color="#f87171", alpha=0.85)

    ax2 = ax1.twinx()
    line = ax2.plot(list(x), remainders, "o-", color="#60a5fa", linewidth=2, markersize=6, label="잔여(만원)")

    for bar in bars_income:
        h = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width() / 2, h + 2, f"{h:,.0f}", ha="center", va="bottom", fontsize=7, color="#e2e8f0")
    for bar in bars_expense:
        h = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width() / 2, h + 2, f"{h:,.0f}", ha="center", va="bottom", fontsize=7, color="#e2e8f0")

    for xi, r in zip(x, remainders):
        color = "#4ade80" if r >= 0 else "#f87171"
        ax2.annotate(f"{r:+,.0f}", (xi, r), textcoords="offset points", xytext=(0, 8),
                     ha="center", fontsize=7, color=color)

    ax1.set_xticks(list(x))
    ax1.set_xticklabels(labels, color="#e2e8f0")
    ax1.set_ylabel("만원", color="#e2e8f0")
    ax2.set_ylabel("잔여 (만원)", color="#60a5fa")
    ax1.tick_params(colors="#e2e8f0")
    ax2.tick_params(colors="#60a5fa")
    for spine in ax1.spines.values():
        spine.set_edgecolor("#4a4a6a")
    for spine in ax2.spines.values():
        spine.set_edgecolor("#4a4a6a")

    handles = [bars_income, bars_expense, line[0]]
    labels_legend = ["수입", "지출", "잔여"]
    ax1.legend(handles, labels_legend, loc="upper left", facecolor="#2e2e4e", labelcolor="#e2e8f0", fontsize=8)

    ax1.set_title("월별 수입/지출 추이", color="#e2e8f0", fontsize=13, pad=12)
    ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:+,.0f}"))
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ── 엑셀 생성 ─────────────────────────────────────────────────────────────────

def generate_excel(months: int = 6) -> bytes:
    """최근 N개월 가계부 Excel bytes 반환."""
    data = _build_monthly_data(months)
    if not data:
        raise ValueError("스냅샷 데이터 없음")

    wb = openpyxl.Workbook()

    # ── Sheet 1: 월별 요약 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "월별 요약"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    pos_fill = PatternFill("solid", fgColor="E2EFDA")
    neg_fill = PatternFill("solid", fgColor="FCE4D6")

    headers = ["월", "총수입(원)", "월급(원)", "추가수입(원)", "고정지출(원)", "카드지출(원)", "총지출(원)", "잔여(원)"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(data, 2):
        ws1.cell(row=row_idx, column=1, value=d["month"])
        ws1.cell(row=row_idx, column=2, value=d["total_income"])
        ws1.cell(row=row_idx, column=3, value=d["salary"])
        ws1.cell(row=row_idx, column=4, value=d["extra_income"])
        ws1.cell(row=row_idx, column=5, value=d["fixed_expense"])
        ws1.cell(row=row_idx, column=6, value=d["card_spend"])
        ws1.cell(row=row_idx, column=7, value=d["total_expense"])
        rem_cell = ws1.cell(row=row_idx, column=8, value=d["remainder"])
        rem_cell.fill = pos_fill if d["remainder"] >= 0 else neg_fill
        for col in range(2, 9):
            ws1.cell(row=row_idx, column=col).number_format = "#,##0"
            ws1.cell(row=row_idx, column=col).alignment = Alignment(horizontal="right")

    col_widths = [10, 16, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 2: 카드사별 지출 ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("카드사별 지출")
    all_cards = sorted({card for d in data for card in d["card_breakdown"].keys()})
    ws2_headers = ["월"] + all_cards + ["합계"]
    for col, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(data, 2):
        ws2.cell(row=row_idx, column=1, value=d["month"])
        for col_idx, card in enumerate(all_cards, 2):
            amount = d["card_breakdown"].get(card, 0)
            cell = ws2.cell(row=row_idx, column=col_idx, value=amount)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        total_cell = ws2.cell(row=row_idx, column=len(all_cards) + 2, value=d["card_spend"])
        total_cell.number_format = "#,##0"
        total_cell.font = Font(bold=True)

    for i in range(1, len(ws2_headers) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
